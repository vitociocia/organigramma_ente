import openpyxl
from django.http import HttpResponse
from django.urls import path
from django.contrib import admin
from django.contrib.admin import SimpleListFilter
from django.db import models
from django.db.models import Q
from django.utils.html import format_html
from django.utils import timezone  # <-- AGGIUNTO

from .models import (
    Qualifica, Responsabile, Struttura, Livello, StrutturaResponsabile
)


# -------------------------
# Inline storico assegnazioni
# -------------------------
class StrutturaResponsabileInline(admin.TabularInline):
    model = StrutturaResponsabile
    extra = 0
    fields = ('responsabile', 'data_inizio', 'data_fine')
    autocomplete_fields = ('responsabile',)


# -------------------------
# Filtri custom
# -------------------------
class AttivaOggiFilter(SimpleListFilter):
    title = "Attiva oggi"
    parameter_name = "attiva_oggi"

    def lookups(self, request, model_admin):
        return (("yes", "Sì"), ("no", "No"))

    def queryset(self, request, qs):
        today = timezone.now().date()  # usa l'import aggiunto
        active = Q(data_inizio__lte=today) & (Q(data_fine__isnull=True) | Q(data_fine__gte=today))
        if self.value() == "yes":
            return qs.filter(active)
        if self.value() == "no":
            return qs.exclude(active)
        return qs


# -------------------------
# Admin Struttura (UNICA registrazione)
# -------------------------
@admin.register(Struttura)
class StrutturaAdmin(admin.ModelAdmin):
    # elenco
    list_display = (
        'nome', 'codice', 'livello', 'responsabile', 'struttura_padre',
        'data_inizio', 'data_fine', 'attiva_oggi'
    )
    list_filter = ('livello', 'responsabile', 'data_inizio', 'data_fine', AttivaOggiFilter)
    search_fields = (
        'nome', 'codice', 'responsabile__nome', 'responsabile__cognome',
        'responsabile__qualifica__titolo'
    )
    ordering = ('codice',)
    autocomplete_fields = ('livello', 'responsabile', 'struttura_padre')
    readonly_fields = ('codice',)
    inlines = [StrutturaResponsabileInline]

    fieldsets = (
        ('Informazioni principali', {
            'fields': ('nome', 'livello', 'struttura_padre', 'responsabile', 'url')
        }),
        ('Periodo di attività', {
            'fields': ('data_inizio', 'data_fine'),
            'description': 'Imposta l’intervallo di validità della struttura.'
        }),
        ('Codici', {
            'fields': ('num_ode', 'num_eng', 'codice'),
            'description': 'Il codice gerarchico è generato automaticamente.'
        }),
    )

    def attiva_oggi(self, obj):
        return obj.is_active()
    attiva_oggi.boolean = True              # <-- icona boolean nativa (✓/✗)
    attiva_oggi.short_description = "Attiva oggi"

    def save_model(self, request, obj, form, change):
        # responsabile precedente (se esiste)
        prev_resp_id = None
        if change:
            prev_resp_id = Struttura.objects.filter(pk=obj.pk).values_list('responsabile_id', flat=True).first()

        super().save_model(request, obj, form, change)

        # sincronizza storico solo se creazione o cambio FK
        if not change or (prev_resp_id != obj.responsabile_id):
            obj.sync_assignment_from_fk(effective_date=timezone.now().date())

    # ----- ACTION: Esporta selezionate in Excel -----
    actions = ["export_selected_excel"]

    def export_selected_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Strutture'
        headers = ["ID", "Nome", "Codice", "Livello", "Data Inizio", "Data Fine",
                   "Resp. Nome", "Resp. Cognome", "Qualifica"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        qs = queryset.select_related('livello', 'responsabile', 'responsabile__qualifica')
        for s in qs:
            r = s.responsabile
            qual = r.qualifica.titolo if r and r.qualifica else "Non Assegnato"
            ws.append([
                s.id, s.nome, s.codice, str(s.livello), s.data_inizio, s.data_fine,
                r.nome if r else "Non Assegnato",
                r.cognome if r else "Non Assegnato",
                qual
            ])

        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename=strutture_selezionate.xlsx'
        wb.save(resp)
        return resp

    export_selected_excel.short_description = "Esporta selezionate in Excel"

    # ----- (opzionale) vista globale export -----
    def export_excel_view(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Strutture'
        headers = ["ID", "Nome", "Codice", "Livello", "Data Inizio", "Data Fine",
                   "Resp. Nome", "Resp. Cognome", "Qualifica"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for s in Struttura.objects.select_related('livello', 'responsabile', 'responsabile__qualifica'):
            r = s.responsabile
            qual = r.qualifica.titolo if r and r.qualifica else "Non Assegnato"
            ws.append([
                s.id, s.nome, s.codice, str(s.livello), s.data_inizio, s.data_fine,
                r.nome if r else "Non Assegnato",
                r.cognome if r else "Non Assegnato",
                qual
            ])

        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename=strutture.xlsx'
        wb.save(resp)
        return resp

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export_excel/', self.admin_site.admin_view(self.export_excel_view), name='export_excel'),
        ]
        return custom_urls + urls


# -------------------------
# Altri ModelAdmin
# -------------------------
@admin.register(Qualifica)
class QualificaAdmin(admin.ModelAdmin):
    list_display = ('titolo', 'dirigente')
    list_filter = ('dirigente',)
    search_fields = ('titolo',)


@admin.register(Responsabile)
class ResponsabileAdmin(admin.ModelAdmin):
    list_display = ('nome', 'cognome', 'qualifica', 'in_carica', 'data_inizio', 'data_fine')
    list_filter = ('qualifica', 'in_carica', 'data_inizio', 'data_fine')
    search_fields = ('nome', 'cognome', 'qualifica__titolo')


@admin.register(Livello)
class LivelloAdmin(admin.ModelAdmin):
    list_display = ('nome', 'ordine', 'can_be_root')
    list_editable = ('ordine', 'can_be_root')
    filter_horizontal = ('allowed_parents',)
    search_fields = ('nome', 'descrizione')
    ordering = ('ordine', 'nome')

admin.site.register(StrutturaResponsabile)
